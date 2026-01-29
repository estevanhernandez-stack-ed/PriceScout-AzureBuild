"""
Daily Lineup Mode - Print-ready theater film schedules
Generates printable daily lineups for individual theaters with blank columns for manual theater numbers.
"""

import streamlit as st
import pandas as pd
import re
from datetime import datetime, date, timedelta, time as dt_time
from app import db_adapter as database
from app.utils import run_async_in_thread


def parse_showtime_for_sort(showtime_str):
    """
    Parse showtime string to a time object for proper chronological sorting.

    Handles formats: HH:MM, H:MM, HH:MM:SS, H:MM:SS
    Returns a time object that sorts correctly.
    """
    if not showtime_str:
        return dt_time(23, 59, 59)  # Put empty times at the end

    try:
        showtime_str = str(showtime_str).strip()

        # Try different formats
        # Include formats with and without space before AM/PM (e.g., "10:00AM" vs "10:00 AM")
        for fmt in ['%H:%M:%S', '%H:%M', '%I:%M %p', '%I:%M:%S %p', '%I:%M%p', '%I:%M:%S%p']:
            try:
                return datetime.strptime(showtime_str, fmt).time()
            except ValueError:
                continue

        # Fallback: try to extract hours and minutes manually
        parts = showtime_str.replace(' ', ':').split(':')
        if len(parts) >= 2:
            hour = int(parts[0])
            minute = int(parts[1])
            return dt_time(hour % 24, minute % 60)

        return dt_time(23, 59, 59)  # Default to end of day
    except:
        return dt_time(23, 59, 59)


def parse_runtime_minutes(runtime_str):
    """
    Parse runtime string to minutes.

    Handles formats: "120 min", "2h 30m", "2:30", "120", "2 hr 30 min"
    Returns integer minutes or None if parsing fails.
    """
    if not runtime_str:
        return None

    try:
        runtime_str = str(runtime_str).strip().lower()

        # Try simple number (assume minutes)
        if runtime_str.isdigit():
            return int(runtime_str)

        # Pattern: "120 min" or "120min"
        match = re.match(r'^(\d+)\s*min', runtime_str)
        if match:
            return int(match.group(1))

        # Pattern: "2h 30m" or "2h30m" or "2 h 30 m"
        match = re.match(r'^(\d+)\s*h(?:r|our)?s?\s*(\d+)?\s*m', runtime_str)
        if match:
            hours = int(match.group(1))
            minutes = int(match.group(2)) if match.group(2) else 0
            return hours * 60 + minutes

        # Pattern: "2:30" (hours:minutes)
        match = re.match(r'^(\d+):(\d+)$', runtime_str)
        if match:
            hours = int(match.group(1))
            minutes = int(match.group(2))
            return hours * 60 + minutes

        # Pattern: just hours "2h" or "2 hours"
        match = re.match(r'^(\d+)\s*h(?:r|our)?s?$', runtime_str)
        if match:
            return int(match.group(1)) * 60

        return None
    except:
        return None


def calculate_outtime(showtime_str, runtime_minutes, use_military_time=False, show_ampm=True):
    """
    Calculate the end time (outtime) given a showtime and runtime.

    Args:
        showtime_str: Showtime in HH:MM or HH:MM:SS format
        runtime_minutes: Runtime in minutes
        use_military_time: If True, use 24-hour format
        show_ampm: If True and not military time, show AM/PM

    Returns:
        Formatted outtime string or None if calculation fails
    """
    if not showtime_str or not runtime_minutes:
        return None

    try:
        # Parse the showtime
        time_obj = parse_showtime_for_sort(showtime_str)
        if time_obj == dt_time(23, 59, 59):  # Failed to parse
            return None

        # Create a datetime to do the math
        base_date = datetime(2000, 1, 1, time_obj.hour, time_obj.minute, time_obj.second)
        end_datetime = base_date + timedelta(minutes=runtime_minutes)

        # Format the outtime using the same format as showtime
        end_time = end_datetime.time()
        if use_military_time:
            return f"{end_time.hour:02d}:{end_time.minute:02d}"
        elif show_ampm:
            hour = end_time.hour % 12
            if hour == 0:
                hour = 12
            period = 'AM' if end_time.hour < 12 else 'PM'
            return f"{hour}:{end_time.minute:02d} {period}"
        else:
            hour = end_time.hour % 12
            if hour == 0:
                hour = 12
            return f"{hour}:{end_time.minute:02d}"
    except:
        return None


def compact_film_title(title, remove_year=True, remove_articles=False, max_words=None):
    """
    Make film titles more compact for narrow column display.

    Args:
        title: Original film title
        remove_year: Remove bracketed year like "(2024)" from end
        remove_articles: Remove leading articles (The, A, An)
        max_words: Limit title to first N words (None = no limit)

    Returns:
        Compacted title string
    """
    if not title:
        return title

    result = title.strip()

    # Remove bracketed year at end (e.g., "(2024)", "(2025)")
    if remove_year:
        result = re.sub(r'\s*\(\d{4}\)\s*$', '', result)

    # Optionally remove leading articles
    if remove_articles:
        result = re.sub(r'^(The|A|An)\s+', '', result, flags=re.IGNORECASE)

    # Limit to max words if specified
    if max_words and max_words > 0:
        words = result.split()
        if len(words) > max_words:
            result = ' '.join(words[:max_words])

    return result.strip()


def render_daily_lineup_mode(cache_data, selected_company):
    """
    Main render function for Daily Lineup Mode.
    Allows theater staff to scrape and generate print-ready daily film schedules.
    """
    st.title("📋 Daily Lineup - Theater Schedule")
    st.info(
        "Scrape your theater's showtimes and generate a print-ready daily film lineup. "
        "Perfect for posting schedules or distribution to staff."
    )

    # Get Marcus and Movie Tavern theaters from cache
    all_theaters = []
    if cache_data:
        for market_name, market_data in cache_data.get('markets', {}).items():
            all_theaters.extend(market_data.get('theaters', []))

    # Filter to only Marcus and Movie Tavern theaters
    company_theaters = [t for t in all_theaters if t.get('company') in ['Marcus', 'Movie Tavern']]
    theater_names = sorted([t['name'] for t in company_theaters if 'name' in t])

    if not theater_names:
        st.warning("No Marcus or Movie Tavern theaters found in cache. Please build the theater cache first.")
        return

    # Theater selection
    st.subheader("Select Theater")

    # Check if user has a default theater (future feature)
    default_theater_index = 0
    if 'user_default_theater' in st.session_state:
        try:
            default_theater_index = theater_names.index(st.session_state.user_default_theater)
        except ValueError:
            pass

    selected_theater = st.selectbox(
        "Theater",
        options=theater_names,
        index=default_theater_index,
        help="Select the theater to generate a lineup for"
    )

    # Get theater object
    selected_theater_obj = next((t for t in company_theaters if t['name'] == selected_theater), None)

    if not selected_theater_obj:
        st.error("Theater not found in cache.")
        return

    st.divider()

    # Date picker
    st.subheader("Select Date")

    today = date.today()
    selected_date_obj = st.date_input(
        "Date",
        value=today,
        min_value=today,
        max_value=today + timedelta(days=30),
        help="Select the date to scrape and generate the lineup for"
    )

    selected_date = selected_date_obj.strftime('%Y-%m-%d')

    st.divider()

    # Display options
    st.subheader("Display Options")

    # Title options row
    col_opt1, col_opt2, col_opt3, col_opt4 = st.columns(4)

    with col_opt1:
        compact_titles = st.checkbox(
            "Compact Titles",
            value=True,
            help="Remove year from titles (e.g., 'Wicked (2024)' → 'Wicked') for narrower columns"
        )

    with col_opt2:
        remove_articles = st.checkbox(
            "Remove Leading Articles",
            value=False,
            help="Remove 'The', 'A', 'An' from start of titles (e.g., 'The Wild Robot' → 'Wild Robot')"
        )

    with col_opt3:
        max_words = st.selectbox(
            "Max Words",
            options=[0, 2, 3, 4, 5],
            index=2,  # Default to 3 words
            format_func=lambda x: "No limit" if x == 0 else f"{x} words",
            help="Limit titles to first N words (e.g., 'Now You See Me: Now You Don't' → 'Now You See')"
        )

    with col_opt4:
        show_outtime = st.checkbox(
            "Show Out Time",
            value=True,
            help="Calculate and display end time based on film runtime"
        )

    # Time format options row
    col_time1, col_time2, col_time3, col_time4 = st.columns(4)

    with col_time1:
        use_military_time = st.checkbox(
            "24-Hour Time",
            value=False,
            help="Use 24-hour (military) time format (e.g., '14:30' instead of '2:30 PM')"
        )

    with col_time2:
        if not use_military_time:
            show_ampm = st.checkbox(
                "Show AM/PM",
                value=True,
                help="Show AM/PM indicator (e.g., '2:30 PM' vs '2:30')"
            )
        else:
            show_ampm = False  # Not applicable for military time

    st.divider()

    # Scrape and Generate button
    if st.button("🔄 Get Latest Showtimes & Generate Lineup", type="primary", use_container_width=True):
        scrape_and_generate(selected_theater_obj, selected_theater, selected_date, selected_date_obj,
                          compact_titles=compact_titles, remove_articles=remove_articles,
                          max_words=max_words if max_words > 0 else None, show_outtime=show_outtime,
                          use_military_time=use_military_time, show_ampm=show_ampm)
    # Display cached lineup data if it exists (persists after download button clicks)
    elif 'lineup_df' in st.session_state and st.session_state.get('lineup_theater') == selected_theater:
        display_cached_lineup(selected_theater, selected_date_obj, show_outtime)


def display_cached_lineup(theater_name, date_obj, show_outtime=True):
    """Display previously generated lineup from session state"""
    lineup_df = st.session_state.get('lineup_df')
    if lineup_df is None or lineup_df.empty:
        return

    # Display header
    st.success(f"✅ Daily Lineup for {theater_name}")
    st.subheader(f"{date_obj.strftime('%A, %B %d, %Y')}")

    # Build column config for dataframe display
    column_config = {
        'Theater #': st.column_config.TextColumn(
            'Theater #',
            width='small',
            help='Leave blank for manual entry'
        ),
        'Film': st.column_config.TextColumn(
            'Film',
            width='large'
        ),
        'In-Time': st.column_config.TextColumn(
            'In-Time',
            width='small'
        ),
    }

    if show_outtime and 'Out-Time' in lineup_df.columns:
        column_config['Out-Time'] = st.column_config.TextColumn(
            'Out-Time',
            width='small',
            help='Calculated end time based on film runtime'
        )

    # Display the lineup table

    # Display the lineup table and add per-row backfill buttons for missing Out-Time
    for idx, row in lineup_df.iterrows():
        st.write(f"**{row['Film']}** | {row['In-Time']}" + (f" → {row['Out-Time']}" if row.get('Out-Time') else ""))
        if show_outtime and (not row.get('Out-Time')):
            if st.button(f"Backfill '{row['Film']}' runtime", key=f"backfill_{idx}"):
                with st.spinner(f"Fetching runtime for '{row['Film']}' from OMDb..."):
                    # Defensive: strip format tags for OMDb query
                    film_title_for_query = re.sub(r"\s*\[.*?\]$", "", row['Film']).strip()
                    count = database.backfill_film_details_from_fandango_single(film_title_for_query)
                    if count > 0:
                        st.success(f"✅ Enriched '{row['Film']}' with runtime!")
                        st.rerun()
                    else:
                        st.warning(f"Could not find runtime for '{row['Film']}'. Check Data Management mode for manual entry.")

    # Optionally, still show the full dataframe below for download/print
    st.dataframe(
        lineup_df,
        use_container_width=True,
        hide_index=True,
        column_config=column_config
    )

    # Printing instructions
    st.divider()
    st.info(
        "📌 **Printing Instructions:**\n"
        "1. Use your browser's print function (Ctrl+P or Cmd+P)\n"
        "2. Set orientation to 'Landscape' for best results\n"
        "3. Adjust scale if needed to fit all content on one page\n"
        "4. The 'Theater #' column can be filled in by hand after printing"
    )

    # Download options using cached data
    st.subheader("Download Options")

    csv_data = st.session_state.get('csv_data')
    csv_filename = st.session_state.get('csv_filename', 'daily_lineup.csv')
    excel_data = st.session_state.get('excel_data')
    xlsx_filename = st.session_state.get('xlsx_filename', 'daily_lineup.xlsx')

    # Use stable keys
    safe_theater_name = re.sub(r'[^\w\-]', '_', theater_name)
    date_str = st.session_state.get('lineup_date', '')
    download_key_base = f"lineup_{safe_theater_name}_{date_str}"

    col1, col2 = st.columns(2)
    with col1:
        if csv_data:
            st.download_button(
                label="📄 Download CSV",
                data=csv_data,
                file_name=csv_filename,
                mime="text/csv",
                use_container_width=True,
                key=f"{download_key_base}_csv_cached"
            )

    with col2:
        if excel_data:
            st.download_button(
                label="📊 Download Excel",
                data=excel_data,
                file_name=xlsx_filename,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
                key=f"{download_key_base}_xlsx_cached"
            )

    # Summary statistics
    st.divider()
    col1, col2, col3 = st.columns(3)
    with col1:
        unique_films = lineup_df['Film'].nunique()
        st.metric("Total Films", unique_films)
    with col2:
        total_showtimes = len(lineup_df)
        st.metric("Total Showtimes", total_showtimes)
    with col3:
        premium_count = sum(1 for _, row in lineup_df.iterrows() if '[' in str(row['Film']))
        st.metric("Premium Format Shows", premium_count)


def scrape_and_generate(theater_obj, theater_name, date_str, date_obj, compact_titles=True, remove_articles=False, max_words=None, show_outtime=True, use_military_time=False, show_ampm=True):
    """Scrape showtimes for a single theater for one date and generate lineup"""
    from app.scraper import Scraper

    theater_url = theater_obj.get('url', '')

    if not theater_url:
        st.error(f"No URL found for {theater_name}")
        return

    st.subheader(f"🔄 Scraping {theater_name} for {date_obj.strftime('%A, %B %d, %Y')}")

    with st.spinner(f"Fetching latest showtimes..."):
        # Initialize scraper
        scout = Scraper()

        # Scrape this theater for this date
        thread, result_func = run_async_in_thread(
            scout.get_all_showings_for_theaters,
            [theater_obj],
            date_str
        )
        thread.join()
        status, result, _, _ = result_func()

        if status == 'success' and result:
            # Save to database using upsert_showings
            database.upsert_showings(result, date_str)

            # Auto-backfill film metadata for any new films discovered
            with st.spinner("Fetching film details from OMDb (runtime, rating, poster)..."):
                try:
                    count = database.backfill_film_details_from_fandango()
                    if count > 0:
                        st.info(f"📊 Enriched {count} film(s) with metadata")
                except Exception as e:
                    st.warning(f"Note: Could not auto-fetch some film details. You can manually backfill in Data Management mode.")

            # Count total showings
            total_showings = sum(len(showings) for showings in result.values())
            st.success(f"✅ Successfully scraped {total_showings} showtimes")
        else:
            st.error(f"Failed to scrape showtimes for {date_str}")
            return

    # Generate the lineup
    st.divider()
    generate_daily_lineup(theater_name, date_str, date_obj, compact_titles=compact_titles,
                         remove_articles=remove_articles, max_words=max_words, show_outtime=show_outtime,
                         use_military_time=use_military_time, show_ampm=show_ampm)


def generate_daily_lineup(theater_name, date_str, date_obj, compact_titles=True, remove_articles=False, max_words=None, show_outtime=True, use_military_time=False, show_ampm=True):
    """Generate and display the daily lineup"""

    # Query showings for this theater and date using SQLAlchemy
    from app.db_adapter import get_session, Showing, Film, config
    from sqlalchemy.orm import aliased
    from datetime import datetime as dt

    # Convert date_str to date object if needed
    if isinstance(date_str, str):
        play_date = dt.strptime(date_str, '%Y-%m-%d').date()
    else:
        play_date = date_str

    with get_session() as session:
        company_id = getattr(config, 'CURRENT_COMPANY_ID', None)

        # Query showings with optional film runtime for outtime calculation
        query = session.query(
            Showing.film_title,
            Showing.showtime,
            Showing.format,
            Showing.daypart,
            Film.runtime
        ).outerjoin(
            Film,
            (Showing.film_title == Film.film_title) & (Showing.company_id == Film.company_id)
        )

        if company_id:
            query = query.filter(Showing.company_id == company_id)

        query = query.filter(
            Showing.theater_name == theater_name,
            Showing.play_date == play_date
        ).order_by(Showing.showtime, Showing.film_title)

        results = query.all()

        if not results:
            st.warning(f"No showtimes found for {theater_name} on {date_str}")
            return

        # Convert to DataFrame
        df = pd.DataFrame(
            results,
            columns=['film_title', 'showtime', 'format', 'daypart', 'runtime']
        )


    # Sort by showtime properly (not alphabetically)
    # This fixes the issue where "10:00" would appear before "9:00"
    df['_sort_time'] = df['showtime'].apply(parse_showtime_for_sort)
    df = df.sort_values(by=['_sort_time', 'film_title']).drop(columns=['_sort_time'])

    # Process the data - create one row per showtime (chronological)
    lineup_data = []

    for _, row in df.iterrows():
        # Format showtime (remove seconds if present)
        formatted_time = format_showtime(row['showtime'], use_military_time=use_military_time, show_ampm=show_ampm)

        # Calculate outtime if runtime is available and option is enabled
        outtime = None
        if show_outtime and row.get('runtime'):
            runtime_mins = parse_runtime_minutes(row['runtime'])
            if runtime_mins:
                outtime = calculate_outtime(row['showtime'], runtime_mins, use_military_time=use_military_time, show_ampm=show_ampm)

        # Get format indicator for this specific showing (returns 'Standard' for regular shows)
        format_indicator = get_format_indicators([row['format']])

        # Apply title compacting if enabled
        film_title = row['film_title']
        if compact_titles or remove_articles or max_words:
            film_title = compact_film_title(film_title, remove_year=compact_titles, remove_articles=remove_articles, max_words=max_words)

        # Append format indicator to film name if it's a premium format (not Standard)
        if format_indicator and format_indicator != 'Standard':
            film_title = f"{film_title} [{format_indicator}]"

        row_data = {
            'Theater #': '',  # Blank column for manual entry
            'Film': film_title,
            'In-Time': formatted_time,
        }

        # Add Out-Time column if enabled
        if show_outtime:
            row_data['Out-Time'] = outtime if outtime else ''

        lineup_data.append(row_data)

    # Create DataFrame for display
    lineup_df = pd.DataFrame(lineup_data)

    # Store in session state for persistent downloads
    st.session_state['lineup_df'] = lineup_df
    st.session_state['lineup_theater'] = theater_name
    st.session_state['lineup_date'] = date_str
    st.session_state['lineup_date_obj'] = date_obj

    # Display header
    st.success(f"✅ Daily Lineup Generated for {theater_name}")
    st.subheader(f"{date_obj.strftime('%A, %B %d, %Y')}")

    # Build column config for dataframe display
    column_config = {
        'Theater #': st.column_config.TextColumn(
            'Theater #',
            width='small',
            help='Leave blank for manual entry'
        ),
        'Film': st.column_config.TextColumn(
            'Film',
            width='large'
        ),
        'In-Time': st.column_config.TextColumn(
            'In-Time',
            width='small'
        ),
    }

    # Add Out-Time column config if showing outtime
    if show_outtime:
        column_config['Out-Time'] = st.column_config.TextColumn(
            'Out-Time',
            width='small',
            help='Calculated end time based on film runtime'
        )

    # Display the lineup table
    st.dataframe(
        lineup_df,
        use_container_width=True,
        hide_index=True,
        column_config=column_config
    )

    # Manual backfill button for missing runtime data
    if 'Out-Time' in lineup_df.columns and (lineup_df['Out-Time'].isna().any() or (lineup_df['Out-Time'] == '').any()):
        st.warning("⚠️ Some films are missing runtime data. Click below to fetch missing details.")
        if st.button("🔄 Backfill Missing Film Details", use_container_width=True):
            with st.spinner("Fetching film details from OMDb..."):
                try:
                    count = database.backfill_film_details_from_fandango()
                    if count > 0:
                        st.success(f"✅ Successfully enriched {count} film(s) with metadata!")
                        st.info("🔃 Refreshing lineup with updated data...")
                        # Regenerate lineup with updated data
                        new_lineup_df = database.generate_daily_lineup(theater_name, date_str)
                        if new_lineup_df is not None and not new_lineup_df.empty:
                            st.session_state['lineup_df'] = new_lineup_df
                            st.rerun()
                        else:
                            st.warning("Could not refresh lineup. Please regenerate manually.")
                    else:
                        st.info("No new film details were found. Films may need manual matching in Data Management mode.")
                except Exception as e:
                    st.error(f"Error during backfill: {str(e)}")
                    st.info("💡 Tip: For films that can't be auto-matched, use the 'Film Details' section in Data Management mode.")

    # Print instructions
    st.divider()
    st.info(
        "📌 **Printing Instructions:**\n"
        "1. Use your browser's print function (Ctrl+P or Cmd+P)\n"
        "2. Set orientation to 'Landscape' for best results\n"
        "3. Adjust scale if needed to fit all content on one page\n"
        "4. The 'Theater #' column can be filled in by hand after printing"
    )

    # Download options
    st.subheader("Download Options")

    # Sanitize theater name for filename
    safe_theater_name = re.sub(r'[^\w\-]', '_', theater_name)
    csv_filename = f"daily_lineup_{safe_theater_name}_{date_str}.csv"
    xlsx_filename = f"daily_lineup_{safe_theater_name}_{date_str}.xlsx"

    # Pre-generate ALL download data and store in session state
    csv_data = lineup_df.to_csv(index=False).encode('utf-8')
    st.session_state['csv_data'] = csv_data
    st.session_state['csv_filename'] = csv_filename

    # Pre-generate Excel with formatting
    try:
        from io import BytesIO
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            lineup_df.to_excel(writer, index=False, sheet_name='Daily Lineup', startrow=2)

            workbook = writer.book
            worksheet = writer.sheets['Daily Lineup']

            worksheet['A1'] = theater_name
            worksheet['A1'].font = Font(size=14, bold=True)
            worksheet['A2'] = date_obj.strftime('%A, %B %d, %Y')
            worksheet['A2'].font = Font(size=12, bold=True)

            header_fill = PatternFill(start_color='8b0e04', end_color='8b0e04', fill_type='solid')
            header_font = Font(color='FFFFFF', bold=True, size=11)

            for col_num, column in enumerate(lineup_df.columns, 1):
                cell = worksheet.cell(row=3, column=col_num)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')

            worksheet.column_dimensions['A'].width = 12
            worksheet.column_dimensions['B'].width = 50
            worksheet.column_dimensions['C'].width = 12
            if 'Out-Time' in lineup_df.columns:
                worksheet.column_dimensions['D'].width = 12

            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            num_cols = len(lineup_df.columns)
            for row in worksheet.iter_rows(min_row=3, max_row=len(lineup_df) + 3, min_col=1, max_col=num_cols):
                for cell in row:
                    cell.border = thin_border
                    cell.alignment = Alignment(vertical='top', wrap_text=True)

            light_fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
            for row_num in range(4, len(lineup_df) + 4, 2):
                for col_num in range(1, num_cols + 1):
                    worksheet.cell(row=row_num, column=col_num).fill = light_fill

        excel_data = output.getvalue()
        st.session_state['excel_data'] = excel_data
        st.session_state['xlsx_filename'] = xlsx_filename
        excel_ready = True
    except ImportError:
        st.caption("Excel export requires openpyxl package")
        excel_ready = False
    except Exception as e:
        st.error(f"Excel export failed: {str(e)}")
        excel_ready = False

    # Display download buttons using session state data
    # Use stable keys to prevent download invalidation on reruns
    download_key_base = f"lineup_{safe_theater_name}_{date_str}"

    col1, col2 = st.columns(2)
    with col1:
        st.download_button(
            label="📄 Download CSV",
            data=csv_data,
            file_name=csv_filename,
            mime="text/csv",
            use_container_width=True,
            key=f"{download_key_base}_csv"
        )

    with col2:
        if excel_ready:
            st.download_button(
                label="📊 Download Excel",
                data=excel_data,
                file_name=xlsx_filename,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
                key=f"{download_key_base}_xlsx"
            )

    # Summary statistics
    st.divider()
    col1, col2, col3 = st.columns(3)
    with col1:
        unique_films = lineup_df['Film'].nunique()
        st.metric("Total Films", unique_films)
    with col2:
        total_showtimes = len(lineup_df)
        st.metric("Total Showtimes", total_showtimes)
    with col3:
        # Count premium formats (indicated by brackets in film name like "[3D]", "[PLF]")
        premium_count = sum(1 for _, row in lineup_df.iterrows() if '[' in str(row['Film']))
        st.metric("Premium Format Shows", premium_count)


def format_showtime(showtime_str, use_military_time=False, show_ampm=True):
    """Format showtime string to be more readable

    Args:
        showtime_str: Time string to format
        use_military_time: If True, use 24-hour format (e.g., '14:30')
        show_ampm: If True and not military time, show AM/PM (e.g., '2:30 PM' vs '2:30')
    """
    try:
        # First, parse the time into a time object
        time_obj = parse_showtime_for_sort(showtime_str)

        if time_obj == dt_time(23, 59, 59):  # Failed to parse
            return showtime_str

        # Format based on options
        if use_military_time:
            # 24-hour format: 14:30
            return f"{time_obj.hour:02d}:{time_obj.minute:02d}"
        elif show_ampm:
            # 12-hour with AM/PM: 2:30 PM
            hour = time_obj.hour % 12
            if hour == 0:
                hour = 12
            period = 'AM' if time_obj.hour < 12 else 'PM'
            return f"{hour}:{time_obj.minute:02d} {period}"
        else:
            # 12-hour without AM/PM: 2:30
            hour = time_obj.hour % 12
            if hour == 0:
                hour = 12
            return f"{hour}:{time_obj.minute:02d}"
    except:
        return showtime_str


def get_format_indicators(formats):
    """Convert format codes to readable indicators"""
    indicators = []

    for fmt in formats:
        # Handle None, empty, or whitespace-only values
        if not fmt or (isinstance(fmt, str) and not fmt.strip()):
            continue

        fmt_str = str(fmt).strip()
        fmt_upper = fmt_str.upper()

        # Skip standard 2D formats
        if fmt_upper in ['STANDARD', '2D', 'STANDARD 2D']:
            continue

        # Common format mappings
        if '3D' in fmt_upper:
            indicators.append('3D')
        if 'IMAX' in fmt_upper:
            indicators.append('IMAX')
        if 'ULTRASCREEN' in fmt_upper:
            indicators.append('UltraScreen')
        if 'PLF' in fmt_upper or 'SUPERSCREEN' in fmt_upper or 'PREMIUM' in fmt_upper:
            indicators.append('PLF')
        if 'DFX' in fmt_upper:
            indicators.append('DFX')
        if 'DOLBY' in fmt_upper:
            indicators.append('Dolby')
        if 'XD' in fmt_upper:
            indicators.append('XD')
        if 'RPX' in fmt_upper:
            indicators.append('RPX')
        if 'DBOX' in fmt_upper or 'D-BOX' in fmt_upper:
            indicators.append('D-BOX')

        # If no specific format matched but it's not standard/2D, show the original
        if not indicators and fmt_upper not in ['STANDARD', '2D', 'STANDARD 2D']:
            indicators.append(fmt_str)

    if not indicators:
        return 'Standard'

    # Remove duplicates and join
    return ', '.join(sorted(set(indicators)))
